# 网页爬取

# 库函数导入
import os
import xlrd
from openpyxl import Workbook


idList = []
nameList = []
urlList = []
fatherList = []

def readData(file):
    # 打开excel文件，创建一个workbook对象,book对象也就是fruits.xlsx文件,表含有sheet名
    rbook = xlrd.open_workbook(path+file)
    # sheets方法返回对象列表,[<xlrd.sheet.Sheet object at 0x103f147f0>]
    rbook.sheets()
    # xls默认有3个工作簿,Sheet1,Sheet2,Sheet3
    rsheet = rbook.sheet_by_index(0)  # 取第一个工作簿
    # 循环工作簿的所有行
    for row in rsheet.get_rows():
        id_column = row[0]  # 品名所在的列
        id_value = id_column.value  # 项目名
        if id_value != '编号':  # 排除第一行
            idList.append(id_value)
            # 读取id
            name_column = row[1]
            name_value = name_column.value
            nameList.append(name_value)
            # 读取名称
            url_column = row[2]
            url_value = url_column.value
            urlList.append(url_value)
            # 读取链接
            father_column = row[3]
            father_value = father_column.value
            fatherList.append(father_value)
            # 读取上级

def writeData():
    # 存储数据到xlxs文件,即excel表格
    wb = Workbook()
    if not os.path.exists(path):  # 判断文件夹是否存在
        os.mkdir(path)  # 新建存储文件夹
    filename = path + '中国地区综合' + '.xlsx'  # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'  # 更改工作表的标题
    ws['A1'] = '编号'  # 对表格加入标题
    ws['B1'] = '名称'
    ws['C1'] = '链接'
    ws['D1'] = '上级'

    for row in range(2, len(idList) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=idList[row - 2])
        _ = ws.cell(column=2, row=row, value=nameList[row - 2])
        _ = ws.cell(column=3, row=row, value=urlList[row - 2])
        _ = ws.cell(column=4, row=row, value=fatherList[row - 2])
    wb.save(filename=filename)  # 保存文件
    print("文件已保存")



if __name__ == "__main__":
    path = 'C:/Users/赵宗天/Desktop/认识Python/爬取文档/'
    file1 = '中国省份.xlsx'
    file2 = '中国城市.xlsx'
    file3 = '中国县区三级.xlsx'
    readData(file1)
    readData(file2)
    readData(file3)
    writeData()

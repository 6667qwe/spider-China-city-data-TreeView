# 网页爬取

# 库函数导入
import os
import requests
import time
import re
import xlrd
from openpyxl import Workbook

provinceList = []  # 省级列表
cityList = []
regionList = []  # 区级列表
root = 'C:/Users/赵宗天/Desktop/认识Python/爬取文档/'
chinaList = []

idList = []
nameList = []
urlList = []
fatherList = []


# 网页爬取函数
# 下面加入了num_retries这个参数，请求失败时重试次数，经过测试网络正常一般最多retry一次就能获得结果
def getUrl(url, num_retries=5):
    headers = {
        'User-Agent': "Mozilla/5.0"}
    try:
        r = requests.get(url, headers=headers)
        r.encoding = r.apparent_encoding
        data = r.text
        return data
    except Exception as e:
        if num_retries > 0:
            # 请求失败时,过 10ms 重试
            time.sleep(10)
            print(url)
            print("请求失败, 重试一下!")
            return getUrl(url, num_retries - 1)  # 递归调用
        else:
            print("请求失败!")
            print("错误: %s" % e + " " + url)
            return  # 返回空值，程序运行报错


# 获取省级代码函数
def getCity(url, pid):
    city = []  # 临时列表，存放市数据
    # print(new_url)
    c_Htmltext = getUrl(url, num_retries=5)
    # soup = BeautifulSoup(c_Htmltext, "html.parser")
    lst = re.findall(r'<a href=\'([^\"\']*)[\'\"]>([\s\S]*?)</a>', c_Htmltext)
    # 市级页面
    # for item in soup.find_all(class_='citytr'):
    for item in lst:
        item = list(item)
        # print(item)
        city.append([item[0], item[1], idList[pid]])
        # c[0] 是父级
    # print(city)
    cityList.append(city)
    print(len(cityList))
    # print(cityList)

def sumData():
    for i in range(len(cityList)):
        print(i)
        for j in range(0, len(cityList[i]), 2):
            print(i)
            chinaList.append([cityList[i][j][1], cityList[i][j + 1][1], cityList[i][j][0], cityList[i][j][2]])
    # print(chinaList)



def savaData():
    # 存储数据到xlxs文件,即excel表格
    wb = Workbook()
    if not os.path.exists(root):  # 判断文件夹是否存在
        os.mkdir(root)  # 新建存储文件夹
    filename = root + '中国城市' + '.xlsx'  # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'  # 更改工作表的标题
    ws['A1'] = '编号'  # 对表格加入标题
    ws['B1'] = '名称'
    ws['C1'] = '链接'
    ws['D1'] = '上级'

    for row in range(2, len(chinaList) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=chinaList[row - 2][0])
        _ = ws.cell(column=2, row=row, value=chinaList[row - 2][1])
        _ = ws.cell(column=3, row=row, value=chinaList[row - 2][2])
        _ = ws.cell(column=4, row=row, value=chinaList[row - 2][3])
    wb.save(filename=filename)  # 保存文件
    print("文件已保存")



def main():
    # 获取省级代码函数
    url = "http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2019/"
    # data = getUrl(url)
    url_list = []
    root = 'C:/Users/赵宗天/Desktop/认识Python/爬取文档/文本文档/'
    # 打开excel文件，创建一个workbook对象,book对象也就是fruits.xlsx文件,表含有sheet名
    rbook = xlrd.open_workbook('C:/Users/赵宗天/Desktop/认识Python/爬取文档/中国省份.xlsx')
    # sheets方法返回对象列表,[<xlrd.sheet.Sheet object at 0x103f147f0>]
    rbook.sheets()
    # xls默认有3个工作簿,Sheet1,Sheet2,Sheet3
    rsheet = rbook.sheet_by_index(0)  # 取第一个工作簿
    # 循环工作簿的所有行
    for row in rsheet.get_rows():
        id_column = row[0]  # 品名所在的列
        id_value = id_column.value  # 项目名
        if id_value != '编号':  # 排除第一行
            idList.append(id_value)
            url_column = row[2]
            url_value = url_column.value
            url_list.append(url_value)
    # print(url_list)
    # 循环获取链接
    for i in range(len(url_list)):
        try:
            data_list = []
            new_url = url + str(url_list[i])
            # print(i)
            # print(new_url)
            getCity(new_url, i)
        except:
            print("获取链接异常")
    sumData()
    savaData()


if __name__ == "__main__":
    # print('*')
    main()

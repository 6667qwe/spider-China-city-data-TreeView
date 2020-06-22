# 网页爬取

# 库函数导入
import os
import requests
import time
import re
from openpyxl import Workbook

provinceList = []  # 省级列表
cityList = []
regionList = []  # 区级列表
root = 'C:/Users/赵宗天/Desktop/认识Python/爬取文档/'
chinaList = []


# 网页爬取函数
# 下面加入了num_retries这个参数，请求失败时重试次数，经过测试网络正常一般最多retry一次就能获得结果
def getUrl(url, num_retries=5):
    headers = {
        'User-Agent': "Mozilla/5.0"}
    try:
        r = requests.get(url, headers=headers)
        r.encoding = r.apparent_encoding
        data = r.text
        return data
    except Exception as e:
        if num_retries > 0:
            # 请求失败时,过 10ms 重试
            time.sleep(10)
            print(url)
            print("请求失败, 重试一下!")
            return getUrl(url, num_retries - 1)  # 递归调用
        else:
            print("请求失败!")
            print("错误: %s" % e + " " + url)
            return  # 返回空值，程序运行报错


# 获取省级代码函数
def getProvince(url):
    # url = 'http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2019/index.html'

    p_Htmltext = getUrl(url, num_retries=5)
    # 省级页面
    # print (Htmltext)

    # 源数据格式  <a href='22.html'>吉林省<br/></a>
    lst = re.findall(r'<a href=\'([^\"\']*)[\'\"]>([\s\S]*?)<br/>', p_Htmltext)
    # 结果 ('13.html', '河北省') 列表

    # 转换结果为 代码 名称 链接  ['13', '河北省','13.html']
    for item in lst:
        item = list(item)
        # item.append(item[0].replace('.html',''))
        provinceList.append([item[0].replace('.html', ''), item[1], item[0]])
        # 去掉.html
        chinaList.append([item[0].replace('.html', ''), item[1], item[0], 0])
    print(chinaList)
    return provinceList


def savaData():
    # 存储数据到xlxs文件,即excel表格
    wb = Workbook()
    if not os.path.exists(root):  # 判断文件夹是否存在
        os.mkdir(root)  # 新建存储文件夹
    filename = root + '中国省份' + '.xlsx'  # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'  # 更改工作表的标题
    ws['A1'] = '编号'  # 对表格加入标题
    ws['B1'] = '名称'
    ws['C1'] = '链接'
    ws['D1'] = '上级'

    for row in range(2, len(chinaList) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=chinaList[row - 2][0])
        _ = ws.cell(column=2, row=row, value=chinaList[row - 2][1])
        _ = ws.cell(column=3, row=row, value=chinaList[row - 2][2])
        _ = ws.cell(column=4, row=row, value=chinaList[row - 2][3])
    wb.save(filename=filename)  # 保存文件
    print("文件已保存")


def main():
    # 获取省级代码函数
    url = "http://www.stats.gov.cn/tjsj/tjbz/tjyqhdmhcxhfdm/2019/index.html"
    # data = getUrl(url)
    getProvince(url)
    savaData()

    # 对 data 循环 调用 getProvince(url) 可获得 市级数据


if __name__ == "__main__":
    print('*')
    main()
